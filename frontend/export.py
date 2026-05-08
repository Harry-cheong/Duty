from __future__ import annotations

import calendar
import datetime as dt
from io import BytesIO
from typing import Any

import pandas as pd


def generate_summary(
    availability_df: pd.DataFrame,
    schedule: pd.DataFrame,
    schedule_name_key: str,
    label: str,
) -> pd.DataFrame:
    summarised_df = availability_df.copy().reset_index(drop=True)
    summarised_df = summarised_df.loc[:, ~summarised_df.columns.astype(str).str.startswith(":")]

    schedule_dates = {
        str(row["date"]).strip()
        for _, row in schedule.iterrows()
        if "date" in row and pd.notna(row["date"])
    }
    for date in schedule_dates:
        if date in summarised_df.columns:
            summarised_df[date] = summarised_df[date].astype("string")

    for _, row in schedule.iterrows():
        name = str(row[schedule_name_key]).strip()
        date = str(row["date"]).strip()
        if date not in summarised_df.columns:
            raise ValueError(f"{date} not found")
        summarised_df.loc[summarised_df["Name"].astype(str).str.strip() == name, date] = label

    return summarised_df


def _normalize_name(value: object) -> str:
    return " ".join(str(value or "").strip().split()).upper()


def _ranked_name(personnel_row: pd.Series) -> str:
    rank = str(personnel_row.get("Rank", "")).strip()
    name = str(personnel_row.get("Name", "")).strip()
    if rank and name:
        return f"{rank} {name}"
    return name or rank


def _build_personnel_lookup(personnel_df: pd.DataFrame) -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    for _, row in personnel_df.iterrows():
        normalized_name = _normalize_name(row.get("Name"))
        if not normalized_name:
            continue
        lookup[normalized_name] = {
            "rank": str(row.get("Rank", "")).strip(),
            "name": str(row.get("Name", "")).strip(),
            "ranked_name": _ranked_name(row),
            "phone": str(row.get("Phone", "")).strip(),
            "department": str(row.get("Department", "")).strip(),
        }
    return lookup


def _display_name(name: str, personnel_lookup: dict[str, dict[str, str]]) -> str:
    person = personnel_lookup.get(_normalize_name(name))
    if person:
        return person["ranked_name"]
    return str(name).strip()


def _parse_slot(slot: str) -> dict[str, Any]:
    text = str(slot).strip()
    parts = text.split()
    date_text = parts[0]
    slot_date = dt.datetime.strptime(date_text, "%d/%m/%Y").date()
    suffix = parts[-1] if parts and parts[-1] in {"AM", "PM"} else ""

    detail_label = slot_date.strftime("%d-%m-%y")
    if suffix:
        detail_label = f"{detail_label} ({suffix})"

    return {
        "slot": text,
        "date": slot_date,
        "weekday": slot_date.strftime("%A"),
        "detail_label": detail_label,
        "suffix": suffix,
    }


def _sheet_name_for_month(year: int, month: int) -> str:
    return f"{calendar.month_abbr[month].upper()} {year % 100:02d} Master Duty Overview"


def _forecast_title_for_month(year: int, month: int) -> str:
    return f"{calendar.month_name[month].upper()} {year} DUTY CLERK FORECAST"


def _schedule_lookup(schedule_df: pd.DataFrame) -> dict[str, str]:
    if schedule_df.empty:
        return {}
    return {
        str(row["date"]).strip(): str(row["assigned_clerk"]).strip()
        for _, row in schedule_df.iterrows()
    }


def _reserve_column_labels(reserve_count: int) -> list[str]:
    return [f"R{index}" for index in range(1, reserve_count + 1)]


def _build_master_overview_df(
    availability_df: pd.DataFrame,
    primary_schedule_df: pd.DataFrame,
    reserve_schedule_dfs: list[pd.DataFrame],
    personnel_lookup: dict[str, dict[str, str]],
) -> tuple[pd.DataFrame, list[dict[str, Any]], list[str]]:
    slot_columns = [
        column
        for column in availability_df.columns
        if column not in {"No", "Name", "Availability"}
    ]
    slot_metadata = [_parse_slot(slot) for slot in slot_columns]

    primary_lookup = _schedule_lookup(primary_schedule_df)
    reserve_lookups = [_schedule_lookup(df) for df in reserve_schedule_dfs]
    reserve_labels = _reserve_column_labels(len(reserve_lookups))

    summary_columns = ["Duty Personnel Duty", "Total Points", "Clerk", *reserve_labels]
    rows: list[dict[str, Any]] = []

    for _, availability_row in availability_df.iterrows():
        person_name = str(availability_row["Name"]).strip()
        display_name = _display_name(person_name, personnel_lookup)

        row: dict[str, Any] = {
            "Duty Personnel Duty": display_name,
            "Total Points": 0,
            "Clerk": 0,
        }
        for reserve_label in reserve_labels:
            row[reserve_label] = 0
        for slot in slot_columns:
            row[slot] = ""

        for slot in slot_columns:
            if primary_lookup.get(slot) == person_name:
                row[slot] = "1"
                row["Clerk"] += 1
                row["Total Points"] += 1
                continue

            for reserve_index, reserve_lookup in enumerate(reserve_lookups, start=1):
                if reserve_lookup.get(slot) != person_name:
                    continue
                reserve_label = f"R{reserve_index}"
                row[slot] = reserve_label.lower()
                row[reserve_label] += 1
                row["Total Points"] += 1
                break

        rows.append(row)

    overview_df = pd.DataFrame(rows, columns=[*summary_columns, *slot_columns])
    return overview_df, slot_metadata, reserve_labels


def _build_send_out_df(
    slot_metadata: list[dict[str, Any]],
    primary_schedule_df: pd.DataFrame,
    reserve_schedule_dfs: list[pd.DataFrame],
    personnel_lookup: dict[str, dict[str, str]],
) -> pd.DataFrame:
    primary_lookup = _schedule_lookup(primary_schedule_df)
    reserve_lookups = [_schedule_lookup(df) for df in reserve_schedule_dfs]

    rows: list[dict[str, Any]] = []
    for slot_info in slot_metadata:
        slot = str(slot_info["slot"])
        primary_name = primary_lookup.get(slot, "")
        primary_person = personnel_lookup.get(_normalize_name(primary_name), {})

        row: dict[str, Any] = {
            "DAY": slot_info["weekday"],
            "DATE": slot_info["detail_label"],
            "CLERK": _display_name(primary_name, personnel_lookup) if primary_name else "",
            "HP NO.": primary_person.get("phone", ""),
            "BRANCH": primary_person.get("department", ""),
        }

        for reserve_index, reserve_lookup in enumerate(reserve_lookups, start=1):
            reserve_name = reserve_lookup.get(slot, "")
            reserve_person = personnel_lookup.get(_normalize_name(reserve_name), {})
            standby_label = "STANDBY" if reserve_index == 1 else f"STANDBY {reserve_index}"
            phone_label = "HP NO. " + str(reserve_index) if reserve_index > 1 else "HP NO. STANDBY"
            branch_label = "BRANCH " + str(reserve_index) if reserve_index > 1 else "BRANCH STANDBY"

            row[standby_label] = _display_name(reserve_name, personnel_lookup) if reserve_name else ""
            row[phone_label] = reserve_person.get("phone", "")
            row[branch_label] = reserve_person.get("department", "")

        rows.append(row)

    return pd.DataFrame(rows)


def _build_personnel_list_df(personnel_df: pd.DataFrame) -> pd.DataFrame:
    export_df = personnel_df.copy().reset_index(drop=True)
    export_df.insert(0, "S/NO.", range(1, len(export_df) + 1))
    export_df["Rank & Name"] = export_df.apply(_ranked_name, axis=1)
    export_df = export_df.rename(
        columns={
            "Rank": "RANK",
            "Name": "NAME",
            "Phone": "HP NO.",
            "Department": "BRANCH",
        }
    )
    return export_df[["S/NO.", "RANK", "NAME", "Rank & Name", "HP NO.", "BRANCH"]]


def _points_history_columns(points_df: pd.DataFrame) -> list[str]:
    return [
        column
        for column in points_df.columns
        if column not in {"Name", "Duty", "Obligation", "Projected"}
    ]


def _build_points_sheet_df(
    duty_points_df: pd.DataFrame,
    reserve_points_df: pd.DataFrame,
    month_label: str,
    personnel_lookup: dict[str, dict[str, str]],
) -> pd.DataFrame:
    duty_history = _points_history_columns(duty_points_df)
    reserve_history = _points_history_columns(reserve_points_df)
    ordered_names = list(
        dict.fromkeys(
            [
                *duty_points_df["Name"].astype(str).str.strip().tolist(),
                *reserve_points_df["Name"].astype(str).str.strip().tolist(),
            ]
        )
    )

    duty_lookup = duty_points_df.copy()
    duty_lookup["Name"] = duty_lookup["Name"].astype(str).str.strip()
    duty_lookup = duty_lookup.set_index("Name")

    reserve_lookup = reserve_points_df.copy()
    reserve_lookup["Name"] = reserve_lookup["Name"].astype(str).str.strip()
    reserve_lookup = reserve_lookup.set_index("Name")

    rows: list[dict[str, Any]] = []
    for name in ordered_names:
        row: dict[str, Any] = {
            "Duty Personnel Duty": _display_name(name, personnel_lookup),
        }
        for column in duty_history:
            row[column] = duty_lookup.at[name, column] if name in duty_lookup.index else pd.NA
        row[f"{month_label}_Duty"] = duty_lookup.at[name, "Projected"] if name in duty_lookup.index else pd.NA
        row["Duty Total"] = (
            duty_lookup.at[name, "Duty"] + duty_lookup.at[name, "Projected"]
            if name in duty_lookup.index
            else pd.NA
        )
        row["Duty Obligation"] = duty_lookup.at[name, "Obligation"] if name in duty_lookup.index else pd.NA

        for column in reserve_history:
            row[column] = reserve_lookup.at[name, column] if name in reserve_lookup.index else pd.NA
        row[f"{month_label}_Reserve"] = reserve_lookup.at[name, "Projected"] if name in reserve_lookup.index else pd.NA
        row["Reserve Total"] = (
            reserve_lookup.at[name, "Duty"] + reserve_lookup.at[name, "Projected"]
            if name in reserve_lookup.index
            else pd.NA
        )
        row["Reserve Obligation"] = reserve_lookup.at[name, "Obligation"] if name in reserve_lookup.index else pd.NA
        rows.append(row)

    return pd.DataFrame(rows)


def _auto_fit_worksheet(worksheet: Any) -> None:
    from openpyxl.utils import get_column_letter

    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        if not values:
            continue
        max_length = max(len(value) for value in values)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)


def _apply_table_style(
    worksheet: Any,
    header_row: int,
    first_data_row: int,
    last_data_row: int,
    last_column: int,
    freeze_cell: str,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    subheader_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for row in worksheet.iter_rows(min_row=header_row, max_row=header_row, min_col=1, max_col=last_column):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    if first_data_row - header_row > 1:
        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=first_data_row - 1,
            min_col=1,
            max_col=last_column,
        ):
            for cell in row:
                cell.fill = subheader_fill
                cell.font = subheader_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

    for row in worksheet.iter_rows(min_row=first_data_row, max_row=last_data_row, min_col=1, max_col=last_column):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    worksheet.freeze_panes = freeze_cell
    worksheet.sheet_view.showGridLines = False
    _auto_fit_worksheet(worksheet)


def _write_master_overview_sheet(
    workbook: Any,
    sheet_name: str,
    overview_df: pd.DataFrame,
    slot_metadata: list[dict[str, Any]],
    reserve_labels: list[str],
) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    worksheet = workbook.create_sheet(title=sheet_name[:31])

    fixed_headers = ["Duty Personnel Duty", "Total Points", "Clerk", *reserve_labels]
    slot_start_column = len(fixed_headers) + 1

    for column_index, header in enumerate(fixed_headers, start=1):
        worksheet.cell(row=1, column=column_index, value=header)
        worksheet.merge_cells(
            start_row=1,
            start_column=column_index,
            end_row=2,
            end_column=column_index,
        )

    for offset, slot_info in enumerate(slot_metadata, start=slot_start_column):
        worksheet.cell(row=1, column=offset, value=slot_info["weekday"])
        worksheet.cell(row=2, column=offset, value=slot_info["detail_label"])

    for row_offset, row in enumerate(overview_df.itertuples(index=False), start=3):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_offset, column=column_index, value=value)

    worksheet["A1"].font = Font(bold=True)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    _apply_table_style(
        worksheet=worksheet,
        header_row=1,
        first_data_row=3,
        last_data_row=overview_df.shape[0] + 2,
        last_column=overview_df.shape[1],
        freeze_cell=f"{get_column_letter(slot_start_column)}3",
    )


def _write_dataframe_sheet(workbook: Any, title: str, dataframe: pd.DataFrame, freeze_cell: str = "A2") -> None:
    worksheet = workbook.create_sheet(title=title[:31])
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)

    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    _apply_table_style(
        worksheet=worksheet,
        header_row=1,
        first_data_row=2,
        last_data_row=max(2, dataframe.shape[0] + 1),
        last_column=dataframe.shape[1],
        freeze_cell=freeze_cell,
    )


def _write_send_out_sheet(workbook: Any, title: str, dataframe: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    worksheet = workbook.create_sheet(title="Send Out")
    last_column = dataframe.shape[1]

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    worksheet.cell(row=1, column=1, value=title)
    worksheet.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=14)
    worksheet.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E78")
    worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for column_index, column_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(row=2, column=column_index, value=column_name)

    for row_index, row in enumerate(dataframe.itertuples(index=False), start=3):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    _apply_table_style(
        worksheet=worksheet,
        header_row=2,
        first_data_row=3,
        last_data_row=max(3, dataframe.shape[0] + 2),
        last_column=last_column,
        freeze_cell="A3",
    )


def build_sample_workbook_bytes(
    *,
    year: int,
    month: int,
    personnel_df: pd.DataFrame,
    availability_df: pd.DataFrame,
    primary_schedule_df: pd.DataFrame,
    reserve_schedule_dfs: list[pd.DataFrame],
    duty_points_df: pd.DataFrame,
    reserve_points_df: pd.DataFrame,
    month_label: str,
) -> bytes:
    from openpyxl import Workbook

    personnel_lookup = _build_personnel_lookup(personnel_df)
    overview_df, slot_metadata, reserve_labels = _build_master_overview_df(
        availability_df=availability_df,
        primary_schedule_df=primary_schedule_df,
        reserve_schedule_dfs=reserve_schedule_dfs,
        personnel_lookup=personnel_lookup,
    )
    send_out_df = _build_send_out_df(
        slot_metadata=slot_metadata,
        primary_schedule_df=primary_schedule_df,
        reserve_schedule_dfs=reserve_schedule_dfs,
        personnel_lookup=personnel_lookup,
    )
    personnel_list_df = _build_personnel_list_df(personnel_df)
    points_df = _build_points_sheet_df(
        duty_points_df=duty_points_df,
        reserve_points_df=reserve_points_df,
        month_label=month_label,
        personnel_lookup=personnel_lookup,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_master_overview_sheet(
        workbook=workbook,
        sheet_name=_sheet_name_for_month(year, month),
        overview_df=overview_df,
        slot_metadata=slot_metadata,
        reserve_labels=reserve_labels,
    )
    _write_send_out_sheet(workbook, _forecast_title_for_month(year, month), send_out_df)
    _write_dataframe_sheet(workbook, "points", points_df, freeze_cell="A2")
    _write_dataframe_sheet(workbook, "Personnel List", personnel_list_df, freeze_cell="A2")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
